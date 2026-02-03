from .utils import *
import csv
import openpyxl
from .models import *
from io import BytesIO
from .serializers import *
from core.permission import *
from datetime import datetime
from openpyxl import Workbook
from django.db.models import Q, F
from decimal import Decimal
from rest_framework import filters
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import permissions, generics, views, status
from rest_framework.parsers import MultiPartParser, FormParser

class CustomLimitPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'limit'
    max_page_size = 50


class ProductListView(generics.ListAPIView):
    """
    Multi-vendor product listing.
    Shows products from PartnerProduct with partner-specific prices, sizes, and colors.
    """
    serializer_class = PartnerProductListSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = CustomLimitPagination    

    # enable filters + search
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['product__name', 'product__description', 'product__brand__name']
    filterset_fields = ['product__sub_category__slug', 'product__gender', 'product__brand']

    def get_queryset(self):
        """
        Build the queryset from PartnerProduct for multi-vendor system.
        Apply filters and optional match score sorting.
        """
        # Get only active partner products
        queryset = PartnerProduct.objects.filter(
            is_active=True,
            product__is_active=True
        ).select_related(
            'product', 'product__brand', 'product__sub_category',
            'partner'
        ).prefetch_related('product__images', 'size_quantities__size', 'color')
        
        # --- Brand filter ---
        brand = self.request.query_params.get("brandName")
        if brand:
            queryset = queryset.filter(product__brand__name__iexact=brand)
        else:
            # No brand requested, exclude Imotana by default
            queryset = queryset.exclude(product__brand__name__iexact="imotana")

        # --- Sub category filter ---
        sub = self.request.query_params.get("sub_category")
        if sub:
            queryset = queryset.filter(product__sub_category__slug=sub)

        # --- Gender filter ---
        gender = self.request.query_params.get("gender")
        if gender:
            queryset = queryset.filter(product__gender=gender)

        # --- Size filter ---
        size_id = self.request.query_params.get("size_id")
        if size_id:
            queryset = queryset.filter(size__id=size_id)

        # --- Color filter ---
        color_id = self.request.query_params.get("color_id")
        if color_id:
            queryset = queryset.filter(color_id=color_id)

        # --- Partner filter ---
        partner_id = self.request.query_params.get("partner_id")
        if partner_id:
            queryset = queryset.filter(partner_id=partner_id)

        # --- Match sorting ---
        match = self.request.query_params.get("match")
        scan = FootScan.objects.filter(user=self.request.user).first()

        if match and match.lower() == "true" and scan:
            partner_products = list(queryset)
            partner_products.sort(
                key=lambda pp: pp.product.match_with_scan(scan).get("score", 0),
                reverse=True  # higher score first
            )
            return partner_products

        # Default: order by latest (descending id)
        queryset = queryset.order_by('-id')

        # --- Filter to show only one variant per unique product to "merge" them globally ---
        from django.db.models import OuterRef, Subquery
        first_variant_sq = PartnerProduct.objects.filter(
            product=OuterRef('product'),
            is_active=True,
            online=True,
            product__is_active=True
        ).order_by('price', 'id').values('id')[:1]
        
        queryset = queryset.filter(id=Subquery(first_variant_sq))

        return queryset

    def list(self, request, *args, **kwargs):
        """
        Apply DRF search + filter backends even when get_queryset() returns a list.
        """
        queryset = self.get_queryset()

        # Apply search and filters manually
        if not isinstance(queryset, list):
            queryset = self.filter_queryset(queryset)

        # Handle pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def get_serializer_context(self):
        context = super().get_serializer_context()
        match = self.request.query_params.get("match")
        scan = FootScan.objects.filter(user=self.request.user).first()
        context["scan"] = scan
        context["match"] = match and match.lower() == "true"
        # Pre-fetch favorite IDs to avoid N+1 queries in serializer
        if self.request.user.is_authenticated:
            favorite_ids = Favorite.objects.filter(user=self.request.user).values_list('products__id', flat=True)
            context["favorite_ids"] = set(favorite_ids)
        else:
            context["favorite_ids"] = set()
        return context


class ProductsCountView(views.APIView):
    """Count available partner products (multi-vendor)"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        sub = request.query_params.get("sub_category")
        gender = request.query_params.get("gender")

        queryset = PartnerProduct.objects.filter(
            is_active=True,
            product__is_active=True
        )
        
        if sub:
            queryset = queryset.filter(product__sub_category__slug=sub)
        
        if gender:
            queryset = queryset.filter(product__gender=gender)
        
        count = queryset.count()

        return Response({"count": count}, status=200)


class ProductDetailView(generics.RetrieveAPIView):
    """
    Multi-vendor product detail view.
    Retrieves a specific PartnerProduct with partner-specific price, size, color.
    """
    serializer_class = PartnerProductDetailSerializer
    permission_classes = [permissions.IsAuthenticated]
    # Lookup by PRODUCT ID, not PartnerProduct ID
    lookup_field = 'id'

    def get_object(self):
        # Find ANY active partner product for this product ID to use as entry point
        # The serializer handles global aggregation
        product_id = self.kwargs.get('id')
        obj = PartnerProduct.objects.filter(
            product__id=product_id,
            is_active=True, 
            product__is_active=True
        ).first()
        
        if not obj:
            # Fallback: check if product exists at all to give better 404
            from django.http import Http404
            raise Http404("Product not found or not available.")
            
        return obj

    def get_serializer_context(self):
        context = super().get_serializer_context()
        scan = FootScan.objects.filter(user=self.request.user).first()
        context['scan'] = scan
        # Pre-fetch favorite IDs to avoid N+1 queries in serializer
        if self.request.user.is_authenticated:
            favorite_ids = Favorite.objects.filter(user=self.request.user).values_list('products__id', flat=True)
            context["favorite_ids"] = set(favorite_ids)
        else:
            context["favorite_ids"] = set()
        return context


class FootScanListCreateView(generics.ListCreateAPIView):
    """List all foot scans or create a new one."""
    serializer_class = FootScanSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return FootScan.objects.filter(user=self.request.user)
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class FavoriteUpdateView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        favorite, created = Favorite.objects.get_or_create(user=request.user)
        favorite_ids = favorite.products.values_list('id', flat=True)
        serializer = FavoriteSerializer(favorite, context={'request': request, 'scan': None, 'match': False, 'favorite_ids': set(favorite_ids)})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, *args, **kwargs):
        favorite, created = Favorite.objects.get_or_create(user=request.user)

        product_id = request.data.get('product_id')
        action = request.data.get('action')

        if not product_id or action not in ['add', 'remove']:
            return Response({"error": "Invalid request."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            product = Product.objects.get(id=product_id, is_active=True)
        except (Product.DoesNotExist, ValueError):
            return Response({"error": "Product not found in inventory."}, status=status.HTTP_404_NOT_FOUND)

        if action == 'add':
            favorite.products.add(product)
            message = "Product added to favorites"
        else:
            favorite.products.remove(product)
            message = "Product removed from favorites"
        
        return Response({"message": message}, status=status.HTTP_200_OK)


class SuggestedProductsView(generics.ListAPIView):
    """
    Suggest similar partner products based on a given partner product.
    Multi-vendor system: returns PartnerProducts with partner-specific pricing.
    """
    serializer_class = PartnerProductListSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = CustomLimitPagination

    def get_queryset(self):
        partner_product_id = self.kwargs.get("product_id")

        if not partner_product_id:
            return PartnerProduct.objects.none()

        try:
            partner_product = PartnerProduct.objects.select_related('product').get(
                id=partner_product_id, 
                is_active=True,
                product__is_active=True
            )
            product = partner_product.product
        except PartnerProduct.DoesNotExist:
            return PartnerProduct.objects.none()

        # Get all SizeTable IDs linked to this product via sizes ManyToMany
        size_table_ids = product.sizes.values_list('id', flat=True)

        # Prepare base queryset - find similar PartnerProducts
        queryset = PartnerProduct.objects.filter(
            is_active=True,
            product__is_active=True,
            product__sub_category=product.sub_category,
            product__gender=product.gender,
            product__sizes__id__in=size_table_ids,  # match same size tables
        ).exclude(
            id=partner_product_id
        ).select_related(
            'product', 'product__brand', 'product__sub_category',
            'partner', 'color'
        ).prefetch_related('product__images', 'size_quantities__size').distinct()

        # Add scan-based ranking (if exists)
        scan = FootScan.objects.filter(user=self.request.user).first()

        if scan:
            foot_width_cat = scan.width_category()
            foot_toe_box = scan.toe_box_category()

            partner_products_list = list(queryset)
            partner_products_list.sort(
                key=lambda pp: (
                    abs(getattr(pp.product, "width", 0) - foot_width_cat),
                    0 if getattr(pp.product, "toe_box", None) == foot_toe_box else 1,
                    -pp.product.match_with_scan(scan).get("score", 0),
                )
            )
            return partner_products_list[:20]  # Top 20 matches

        return queryset[:20]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        scan = FootScan.objects.filter(user=self.request.user).first()
        context["scan"] = scan
        context["match"] = True
        # Pre-fetch favorite IDs to optimize serializer
        if self.request.user.is_authenticated:
            favorite_ids = Favorite.objects.filter(user=self.request.user).values_list('products__id', flat=True)
            context["favorite_ids"] = set(favorite_ids)
        else:
            context["favorite_ids"] = set()
        return context

                
class ProductQnAFilterAPIView(views.APIView):
    """
    Filter products by Q&A in multi-vendor system.
    Returns PartnerProducts with partner-specific pricing.
    """
    pagination_class = CustomLimitPagination
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        data = request.data

        if not isinstance(data, dict):
            return Response(
                {"error": "Invalid data format, expected a JSON object."},
                status=status.HTTP_400_BAD_REQUEST
            )

        cat = data.get("sub_category")
        questions_data = data.get("questions", [])

        # If no questions provided, return empty
        if not questions_data:
            return self.empty_paginated_response(request)

        if not cat:
            return self.empty_paginated_response(request)

        # Build OR query across ALL questions
        combined_query = Q()
        
        for question_item in questions_data:
            question_key = question_item.get("question")
            answer_keys = question_item.get("answers", [])
            
            if not question_key or not answer_keys:
                continue
            
            # Find the question object by EXACT key match
            try:
                question_obj = Question.objects.get(key=question_key)
            except Question.DoesNotExist:
                continue
            
            # For this question, build OR query for all answer options
            for answer_key in answer_keys:
                if answer_key and answer_key.strip():
                    combined_query |= Q(
                        product__question_answers__question=question_obj,
                        product__question_answers__answers__key=answer_key
                    )
        
        # If no valid query built, return empty
        if not combined_query:
            return self.empty_paginated_response(request)
        
        # Get PartnerProducts matching the Q&A filter
        queryset = PartnerProduct.objects.filter(
            is_active=True,
            product__is_active=True,
            product__sub_category__slug=cat
        ).filter(combined_query).select_related(
            'product', 'product__brand', 'product__sub_category',
            'partner'
        ).prefetch_related('product__images', 'size_quantities__size', 'color').distinct()
        
        # Final check
        if not queryset.exists():
            return self.empty_paginated_response(request)

        # Get scan for context
        scan = FootScan.objects.filter(user=request.user).first()

        # Pagination and serialization
        paginator = self.pagination_class()
        paginated_products = paginator.paginate_queryset(queryset, request, view=self)
        serializer = PartnerProductListSerializer(
            paginated_products, 
            many=True, 
            context={'request': request, 'scan': scan, 'favorite_ids': set(Favorite.objects.filter(user=request.user).values_list('products__id', flat=True)) if request.user.is_authenticated else set()}
        )
        
        return paginator.get_paginated_response(serializer.data)

    # Helper method for empty response
    def empty_paginated_response(self, request):
        paginator = self.pagination_class()
        empty_queryset = PartnerProduct.objects.none()
        paginated = paginator.paginate_queryset(empty_queryset, request, view=self)
        serializer = PartnerProductListSerializer(paginated, many=True)
        return paginator.get_paginated_response(serializer.data)


class AllProductsForPartnerView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated, IsPartner]
    pagination_class = CustomLimitPagination
    serializer_class = ProductSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name', 'description', 'brand__name']
    filterset_fields = ['sub_category__slug', 'gender', 'brand']

    def get_queryset(self):
        queryset = Product.objects.filter(is_active=True).order_by('-id')
        return queryset


class SingleProductForPartnerView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, IsPartner]
    
    def get(self, request, product_id):
        # 1. Try to find by PartnerProduct ID (specific variant)
        partner_product = PartnerProduct.objects.filter(id=product_id, partner=request.user).first()
        
        if not partner_product:
            # 2. Fallback: Lookup by Product ID (catalogue product)
            queryset = PartnerProduct.objects.filter(product_id=product_id, partner=request.user)
            
            # Allow filtering by color if multiple variants for the same product exist
            color_id = request.query_params.get('color_id')
            if color_id:
                partner_product = queryset.filter(color_id=color_id).first()
            else:
                # Just pick the first one found for this product
                partner_product = queryset.first()

        if partner_product:
            serializer = PartnerProductSerializer(partner_product)
            return Response(serializer.data)
        
        return Response({"error": "Partner product not found"}, status=status.HTTP_404_NOT_FOUND)


class ApprovedPartnerProductView(generics.ListAPIView):
    """View to show all products in the partner's inventory"""
    permission_classes = [permissions.IsAuthenticated, IsPartner]
    # pagination_class = CustomLimitPagination
    serializer_class = PartnerProductSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['product__name', 'product__description', 'product__brand__name','warehouse__name']
    filterset_fields = ['product__sub_category__slug', 'product__gender', 'product__brand','warehouse__id']

    def get_queryset(self):
        # Return all PartnerProduct entries for this partner
        queryset = PartnerProduct.objects.filter(partner=self.request.user).select_related('product', 'product__brand', 'color').prefetch_related('product__images', 'product__images__color')
        
        # --- Warehouse filter ---
        warehouse_id = self.request.query_params.get("warehouse_id")
        if warehouse_id:
            queryset = queryset.filter(warehouse__id=warehouse_id)
            
        return queryset


class ApprovedPartnerProductUpdateView(views.APIView):
    """View to add or remove products from partner's inventory"""
    permission_classes = [permissions.IsAuthenticated, IsPartner]

    def patch(self, request, *args, **kwargs):
        product_id = kwargs.get('product_id')
        print('Product ID :', product_id)
        warehouse_id = request.data.get('warehouse_id', None)
        action = kwargs.get('action')
        eanc = request.data.get('eanc', None)
        buy_price = request.data.get('buy_price', 0)
        
        # Valid actions: add, update, del/remove
        if not product_id or action not in ['add', 'update', 'del', 'remove']:
            return Response({"error": "Invalid request. Provide product_id and valid action (add/update/del)."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Removed is_active=True filter so partners can still manage (e.g., delete) 
            # products even if they are inactive or local-only.
            product = Product.objects.get(id=product_id)
            print('Product :', product,product_id,product.id)
        except Product.DoesNotExist:
            return Response({"error": "Product not found."}, status=status.HTTP_404_NOT_FOUND)

        if action == 'add':

            price = request.data.get('price')
            size_quantities = request.data.get('sizes', [])

            if not price:
                return Response({"error": "Price is required when adding a product."}, status=status.HTTP_400_BAD_REQUEST)
            # Resolve Color
            color_input = request.data.get('color')
            color_obj = None
            if color_input:
                color_input = str(color_input).strip()
                if color_input.isdigit():
                    color_obj = Color.objects.filter(id=int(color_input)).first()
                if not color_obj:
                    color_obj = Color.objects.filter(color__iexact=color_input).first()
                    if not color_obj:
                        color_obj = Color.objects.create(color=color_input, hex_code="#000000")
            
            if not color_obj:
                return Response({"error": "Color is required and not found."}, status=status.HTTP_400_BAD_REQUEST)
            print('Product :', product)
            print('Color :', color_obj)
            # Check if already exists
            if PartnerProduct.objects.filter(partner=request.user, product=product, color=color_obj).exists():
                return Response({"error": "Product already exists in your inventory. Use 'update' action instead."}, status=status.HTTP_400_BAD_REQUEST)

            
            # --- Size Resolution Logic ---
            resolved_sizes = [] # List of {'size_id': id, 'quantity': q}

            if isinstance(size_quantities, dict):
                # Format: {"EU 40": 10}
                for size_str, qty in size_quantities.items():
                    parts = str(size_str).split()
                    s_type, s_value = (parts[0], " ".join(parts[1:])) if len(parts) >= 2 else ("EU", size_str)
                    
                    # 1. Prioritize finding a size that is ALREADY linked to this product (Official Size)
                    size_obj = Size.objects.filter(
                        table__product_images__product=product,
                        type=s_type,
                        value=s_value
                    ).distinct().first()
                    
                    # 2. If not found, try brand tables
                    if not size_obj:
                        size_obj = Size.objects.filter(table__brand=product.brand, type=s_type, value=s_value).first()
                    
                    # 3. If still not found, fallback to global
                    if not size_obj:
                        size_obj = Size.objects.filter(type=s_type, value=s_value).first()
                    
                    # 4. If still doesn't exist anywhere, create local
                    if not size_obj:
                        local_table, _ = SizeTable.objects.get_or_create(brand=product.brand, name="Local Sizes")
                        size_obj = Size.objects.create(table=local_table, type=s_type, value=s_value, insole_min_mm=0, insole_max_mm=0)
                    
                    resolved_sizes.append({'size_id': size_obj.id, 'label': size_str, 'quantity': qty})
            
            elif isinstance(size_quantities, list):
                # Format: [{"size_id": 1, "quantity": 10}]
                resolved_sizes = size_quantities
            else:
                return Response({"error": "Sizes format must be a list or a dictionary."}, status=status.HTTP_400_BAD_REQUEST)

            if not resolved_sizes:
                return Response({"error": "Sizes with quantities are required."}, status=status.HTTP_400_BAD_REQUEST)
            
            color_id = color_obj.id

            # Determine online status: 
            # 1. Must have global image for this color
            # 2. All sizes must be in the official global size tables for this color image
            product_image = ProductImage.objects.filter(product=product, color=color_obj).first()
            has_global_image = product_image is not None
            
            if not has_global_image:
                print(f"DEBUG: Product {product.id} color {color_obj} has no global image. Marking as local.")

            all_sizes_official = True
            if has_global_image:
                # Get all official sizes for this specific product-color image
                official_size_ids = set(Size.objects.filter(
                    table__product_images__product=product
                ).values_list('id', flat=True).distinct())
                
                # Also check sizes specifically linked to this color's image
                color_specific_size_ids = set(Size.objects.filter(
                    table__product_images=product_image
                ).values_list('id', flat=True))
                
                # Merge both sets - a size is official if it's in either the product's tables OR the color-specific tables
                all_official_size_ids = official_size_ids | color_specific_size_ids
                
                for sq_data in resolved_sizes:
                    s_id = sq_data.get('size_id')
                    s_label = sq_data.get('label', f"ID:{s_id}")
                    if s_id not in all_official_size_ids:
                        print(f"DEBUG: Size '{s_label}' (Resolved ID: {s_id}) is not in official size table for Product {product.id} color {color_obj}. Marking as local.")
                        all_sizes_official = False
                        break
            else:
                all_sizes_official = False

            can_be_online = has_global_image and all_sizes_official
            
            # Final online status decision
            requested_online = request.data.get('online', False)
            if requested_online and not can_be_online:
                # User is trying to set online=True but product doesn't meet requirements
                error_reasons = []
                if not has_global_image:
                    error_reasons.append(f"No global image found for color '{color_obj.color}'")
                if not all_sizes_official:
                    # Get the unofficial size details
                    unofficial_sizes = []
                    for sq_data in resolved_sizes:
                        s_id = sq_data.get('size_id')
                        s_label = sq_data.get('label', f"ID:{s_id}")
                        if s_id not in all_official_size_ids:
                            unofficial_sizes.append(s_label)
                    if unofficial_sizes:
                        error_reasons.append(f"Unofficial sizes detected: {', '.join(unofficial_sizes)}")
                    else:
                        error_reasons.append("One or more sizes are not in the official size table")
                
                return Response({
                    "error": "This product cannot be set to online status.",
                    "reasons": error_reasons,
                    "message": "Please ensure the product has a global image and all sizes are from the official size table. The product will be added as local-only."
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Set online status
            if requested_online and not can_be_online:
                online = False
            else:
                # Default to true if possible and no explicit preference, otherwise respect request
                online = requested_online or (can_be_online and request.data.get('online') is None)
            
            # Force False if explicitly requested
            if request.data.get('online') is False:
                online = False

            # Check if already exists for this specific color
            if PartnerProduct.objects.filter(partner=request.user, product=product, color_id=color_id).exists():
                return Response({"error": "This color variant already exists."}, status=status.HTTP_400_BAD_REQUEST)

            partner_product = PartnerProduct.objects.create(
                partner=request.user,
                product=product,
                buy_price=buy_price,
                price=price,
                color_id=color_id,
                is_active=online,
                eanc=eanc,
                online=online,
                local=request.data.get('local', True)
            )

            for sq_data in resolved_sizes:
                PartnerProductSize.objects.create(
                    partner_product=partner_product,
                    size_id=sq_data.get('size_id'),
                    quantity=sq_data.get('quantity', 0)
                )

            if warehouse_id:
                try:
                    warehouse = Warehouse.objects.get(id=warehouse_id, partner=request.user)
                    warehouse.product.add(partner_product)
                except Warehouse.DoesNotExist:
                    pass

            serializer = PartnerProductSerializer(partner_product)
            return Response({"message": "Product variant added to inventory", "data": serializer.data}, status=status.HTTP_201_CREATED)

        elif action == 'update':
            # 1. Try to find the specific variant
            # product_id here could be either the base product ID or the specific PartnerProduct ID
            partner_product = PartnerProduct.objects.filter(id=product_id, partner=request.user).first()
            
            if not partner_product:
                # Resolve color provided in request to find the variant
                color_input = request.data.get('color') or request.data.get('color_id')
                if color_input:
                    color_input = str(color_input).strip()
                    color_obj = None
                    if color_input.isdigit():
                        color_obj = Color.objects.filter(id=int(color_input)).first()
                    
                    if not color_obj:
                        color_obj = Color.objects.filter(color__iexact=color_input).first()
                        if not color_obj:
                            color_obj = Color.objects.create(color=color_input, hex_code="#000000")
                    
                    if color_obj:
                        partner_product = PartnerProduct.objects.filter(
                            partner=request.user, 
                            product_id=product_id, 
                            color=color_obj
                        ).first()
            
            if not partner_product:
                 return Response({"error": "Product variant not found in your inventory."}, status=status.HTTP_404_NOT_FOUND)

            # 2. Update basic fields
            if 'price' in request.data:
                partner_product.price = request.data.get('price')
            
            if 'eanc' in request.data:
                partner_product.eanc = request.data.get('eanc')
            
            if 'local' in request.data:
                partner_product.local = request.data.get('local')

            partner_product.save()

            # 3. Handle sizes update
            if 'sizes' in request.data:
                size_input = request.data.get('sizes')
                resolved_updates = []

                if isinstance(size_input, dict):
                    # Format: {"EU 40": 10}
                    for size_str, qty in size_input.items():
                        parts = str(size_str).split()
                        s_type, s_value = (parts[0], " ".join(parts[1:])) if len(parts) >= 2 else ("EU", size_str)
                        
                        # 1. Prioritize official size linked to this product
                        size_obj = Size.objects.filter(
                            table__product_images__product=partner_product.product,
                            type=s_type,
                            value=s_value
                        ).distinct().first()
                        
                        # 2. Try brand tables
                        if not size_obj:
                            size_obj = Size.objects.filter(table__brand=partner_product.product.brand, type=s_type, value=s_value).first()
                        
                        # 3. Fallback to global
                        if not size_obj:
                            size_obj = Size.objects.filter(type=s_type, value=s_value).first()
                        
                        # 4. Create local if missing
                        if not size_obj:
                            local_table, _ = SizeTable.objects.get_or_create(brand=partner_product.product.brand, name="Local Sizes")
                            size_obj = Size.objects.create(table=local_table, type=s_type, value=s_value, insole_min_mm=0, insole_max_mm=0)
                        
                        resolved_updates.append({'size_id': size_obj.id, 'quantity': qty})
                elif isinstance(size_input, list):
                    resolved_updates = size_input

                if resolved_updates:
                    resolved_map = {}
                    for sq_data in resolved_updates:
                        input_id = sq_data.get('size_id')
                        if input_id is None: continue
                        
                        # Check if it's already a PartnerProductSize ID
                        pps = PartnerProductSize.objects.filter(id=input_id, partner_product=partner_product).first()
                        if pps:
                            resolved_map[input_id] = pps.size_id
                            continue
                            
                        if Size.objects.filter(id=input_id).exists():
                            resolved_map[input_id] = input_id
                            continue
                        
                        return Response({"error": f"Size ID {input_id} not found."}, status=status.HTTP_404_NOT_FOUND)

                    for sq_data in resolved_updates:
                        input_id = sq_data.get('size_id')
                        quantity = sq_data.get('quantity', 0)
                        real_size_id = resolved_map.get(input_id)
                        if real_size_id is not None:
                            print(f"DEBUG: Updating PartnerProductSize: Variant:{partner_product.id}, SizeID:{real_size_id}, NewQty:{quantity}")
                            PartnerProductSize.objects.update_or_create(
                                partner_product=partner_product,
                                size_id=real_size_id,
                                defaults={'quantity': quantity}
                            )

            # 4. Handle Online and Active status validation (After size updates)
            # Must have global image and ALL sizes must be official
            product_image = ProductImage.objects.filter(
                product=partner_product.product, 
                color=partner_product.color
            ).first()
            has_global_image = product_image is not None
            
            all_sizes_official = True
            if has_global_image:
                # Get all official sizes for this product (from ALL images, not just this color)
                # This is because sizes are typically shared across colors for the same product
                official_size_ids = set(Size.objects.filter(
                    table__product_images__product=partner_product.product
                ).values_list('id', flat=True).distinct())
                
                # Also check sizes specifically linked to this color's image
                color_specific_size_ids = set(Size.objects.filter(
                    table__product_images=product_image
                ).values_list('id', flat=True))
                
                # Merge both sets - a size is official if it's in either the product's tables OR the color-specific tables
                all_official_size_ids = official_size_ids | color_specific_size_ids
                
                # Check ALL current sizes for this partner product
                current_sizes = PartnerProductSize.objects.filter(partner_product=partner_product)
                for pps in current_sizes:
                    if pps.size_id not in all_official_size_ids:
                        print(f"DEBUG: Size ID {pps.size_id} for PartnerProduct {partner_product.id} is unofficial. Marking online=False.")
                        all_sizes_official = False
                        break
            else:
                all_sizes_official = False

            can_be_online = has_global_image and all_sizes_official

            # Logic for 'online'
            if 'online' in request.data:
                requested_online = request.data.get('online')
                if requested_online and not can_be_online:
                    # User is trying to set online=True but product doesn't meet requirements
                    error_reasons = []
                    if not has_global_image:
                        error_reasons.append(f"No global image found for color '{partner_product.color.color}'")
                    if not all_sizes_official:
                        # Get the unofficial size details
                        unofficial_sizes = []
                        current_sizes = PartnerProductSize.objects.filter(partner_product=partner_product).select_related('size')
                        for pps in current_sizes:
                            if pps.size_id not in all_official_size_ids:
                                unofficial_sizes.append(f"{pps.size.type} {pps.size.value}")
                        if unofficial_sizes:
                            error_reasons.append(f"Unofficial sizes detected: {', '.join(unofficial_sizes)}")
                        else:
                            error_reasons.append("One or more sizes are not in the official size table")
                    
                    return Response({
                        "error": "This product cannot be set to online status.",
                        "reasons": error_reasons,
                        "message": "Please ensure the product has a global image and all sizes are from the official size table."
                    }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    partner_product.online = requested_online
            else:
                # If sizes changed, we might need to auto-disable online
                if not can_be_online:
                    partner_product.online = False

            # Logic for 'is_active'
            if 'is_active' in request.data:
                requested_active = request.data.get('is_active')
                if requested_active and not (partner_product.online or partner_product.local):
                    partner_product.is_active = False
                elif requested_active and partner_product.online and not can_be_online:
                    partner_product.is_active = False
                else:
                    partner_product.is_active = requested_active
            else:
                # IMPORTANT: Automatically enable is_active if either online or local is True
                partner_product.is_active = (partner_product.online or partner_product.local)
            
            # Update warehouse if provided
            if warehouse_id and Warehouse.objects.filter(id=warehouse_id).exists():
                partner_product.warehouse_id = warehouse_id
            
            # 5. Save changes
            partner_product.save()

            serializer = PartnerProductSerializer(partner_product)
            return Response({"message": "Product variant updated in inventory", "data": serializer.data}, status=status.HTTP_200_OK)

        elif action in ['del', 'remove']:
            # 1. Try to find the specific variant by Direct ID
            partner_product = PartnerProduct.objects.filter(id=product_id, partner=request.user).first()
            print(partner_product,'partner_product')
            if not partner_product:
                # 2. Resolve color provided in request to find the variant
                color_input = request.data.get('color') or request.data.get('color_id')
                if color_input:
                    color_input = str(color_input).strip()
                    color_obj = None
                    if color_input.isdigit():
                        color_obj = Color.objects.filter(id=int(color_input)).first()
                    
                    if not color_obj:
                        color_obj = Color.objects.filter(color__iexact=color_input).first()
                    
                    if color_obj:
                        partner_product = PartnerProduct.objects.filter(
                            partner=request.user, 
                            product_id=product_id, 
                            color=color_obj
                        ).first()
            
            if not partner_product: 
                return Response({"error": "Product variant not found in your inventory."}, status=status.HTTP_404_NOT_FOUND)
            else:
                deleted_count, _ = partner_product.delete()

            if deleted_count == 0:
                return Response({"error": "Product not found in your inventory."}, status=status.HTTP_404_NOT_FOUND)
            return Response({"message": "Product variant removed from inventory"}, status=status.HTTP_200_OK)

        else:
            return Response({"error": "Invalid action."}, status=status.HTTP_400_BAD_REQUEST)


class FileUploadPartnerProductView(views.APIView):
    """
    View to upload and read data from Excel or CSV files for partner products.
    """
    permission_classes = [permissions.IsAuthenticated, IsPartner]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        filename = file.name
        data = []

        if filename.endswith('.csv'):
            try:
                # Read CSV file
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                for row in reader:
                    data.append(row)
            except Exception as e:
                return Response({"error": f"Error reading CSV: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                # Read Excel file
                workbook = openpyxl.load_workbook(file, data_only=True)
                sheet = workbook.active
                
                # Get headers from the first row
                headers = [cell.value for cell in sheet[1]]
                
                # Iterate through rows starting from the second row
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Create a dictionary for each row mapping headers to values
                    row_data = dict(zip(headers, row))
                    data.append(row_data)
            except Exception as e:
                return Response({"error": f"Error reading Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"error": "Unsupported file format. Please upload CSV or XLSX."}, status=status.HTTP_400_BAD_REQUEST)

        return self.process_uploaded_data(request, data)

    def process_uploaded_data(self, request, data):
        partner = request.user
        
        # Summary Tracking
        summary = {
            "total_rows": len(data),
            "matched_products": 0,
            "created_products": 0,
            "new_partner_products": 0,
            "updated_partner_products": 0,
            "size_updates": 0,
            "color_updates": 0,
            "local_only_count": 0,
            "skipped_rows": 0
        }
        skipped_details = []
        
        # Cache for performance
        product_cache = {}
        
        # Get Warehouse if provided
        warehouse_id = request.data.get('warehouse_id')
        warehouse = None
        if warehouse_id:
            try:
                from Others.models import Warehouse
                warehouse = Warehouse.objects.get(id=warehouse_id, partner=partner)
            except:
                pass

        for row in data:
            item_name = row.get('Item name') or row.get('item_name')
            if not item_name:
                summary["skipped_rows"] += 1
                skipped_details.append({"row": row, "reason": "Missing 'Item name'"})
                continue
            
            is_local_only = False
            
            # --- 1. Resolve Product ---
            product = product_cache.get(item_name)
            if not product:
                product = Product.objects.filter(name__iexact=item_name).first()
                if not product:
                    product = Product.objects.filter(name__icontains=item_name).first()
                
                if product:
                    product_cache[item_name] = product
                else:
                    # Product not found -> Create a "Local" product
                    is_local_only = True
                    try:
                        local_brand, _ = Brand.objects.get_or_create(name="Local")
                        local_cat, _ = Category.objects.get_or_create(name="Local", defaults={'slug': 'local-cat'})
                        local_subcat, _ = SubCategory.objects.get_or_create(
                            name="Local", 
                            category=local_cat, 
                            defaults={'slug': 'local-subcat'}
                        )
                        
                        product = Product.objects.create(
                            name=item_name,
                            brand=local_brand,
                            main_category=local_cat,
                            sub_category=local_subcat,
                            description=f"Auto-generated local product for {item_name}",
                            gender="unisex",
                            is_active=False # Keep it inactive for global search maybe?
                        )
                        product_cache[item_name] = product
                        summary["created_products"] += 1
                    except Exception as e:
                        summary["skipped_rows"] += 1
                        skipped_details.append({"item": item_name, "reason": f"Failed to create local product: {str(e)}"})
                        continue
            
            if not is_local_only:
                summary["matched_products"] += 1

            # --- 2. Resolve Color ---
            color_name = row.get('Color name') or row.get('color_name')
            color_id_val = row.get('color_id')
            image_id_val = row.get('image_id')
            
            color_obj = None
            
            # Try to match via existing product images first (Online mode)
            if color_name:
                target_color = Color.objects.filter(color__iexact=str(color_name).strip()).first()
                if target_color and ProductImage.objects.filter(product=product, color=target_color).exists():
                    color_obj = target_color
            
            if not color_obj and image_id_val:
                img = ProductImage.objects.filter(id=image_id_val, product=product).select_related('color').first()
                if img and img.color:
                    color_obj = img.color
            
            # If still not found, color does not match product's global catalog -> mark as local only
            if not color_obj:
                is_local_only = True
                if color_name:
                    color_obj, _ = Color.objects.get_or_create(
                        color=str(color_name).strip(),
                        defaults={'hex_code': '#CCCCCC'} # Default grey for local colors
                    )
                elif color_id_val:
                    color_obj = Color.objects.filter(id=color_id_val).first()

            if not color_obj:
                # Still no color? Pick a default or skip
                summary["skipped_rows"] += 1
                skipped_details.append({"item": item_name, "reason": "No valid color found and no color name provided"})
                continue

            # --- 3. Resolve/Create PartnerProduct ---
            price_str = row.get('Amount Incl. VAT') or row.get('amount_incl_vat') or row.get('Price') or '0'
            price = self.clean_price(price_str)
            
            partner_product, created = PartnerProduct.objects.get_or_create(
                partner=partner,
                product=product,
                color=color_obj,
                defaults={
                    'price': price,
                    'online': not is_local_only,
                    'local': True,
                    'is_active': True
                }
            )
            
            if created:
                summary["new_partner_products"] += 1
            else:
                summary["updated_partner_products"] += 1
                partner_product.price = price
                # If it was already online=False, keep it. If now it's newly local, set it.
                if is_local_only:
                    partner_product.online = False
                partner_product.save()

            if is_local_only:
                summary["local_only_count"] += 1

            # Associate Color - Now already handled via get_or_create
            # partner_product.color.add(color_obj)  # Removed as it's now a FK
            summary["color_updates"] += 1

            if warehouse:
                warehouse.product.add(partner_product)
            
            if row.get('EAN') or row.get('ean'):
                partner_product.eanc = str(row.get('EAN') or row.get('ean')).strip()
                partner_product.save()

            # --- 4. Resolve Sizes (EU, US, etc.) ---
            # Define which size types are relevant for this product's gender
            relevant_types = ['EU']
            if product.gender == 'male':
                relevant_types.append('USM')
            elif product.gender == 'female':
                relevant_types.append('USW')
            else: # unisex or others
                relevant_types.extend(['USM', 'USW'])

            size_cols = [
                ('Size EU', 'EU'),
                ('Size US', 'USM'),
                ('Size US', 'USW'),
                ('size_eu', 'EU'),
                ('size_us', 'USM'),
                ('size_us', 'USW'), # Added for completeness
            ]
            
            quantity = 0
            try:
                quantity_val = row.get('Quantity') or row.get('quantity')
                quantity = int(quantity_val) if quantity_val is not None else 0
            except:
                pass

            processed_columns_in_row = set()
            for col, s_type in size_cols:
                if s_type not in relevant_types:
                    continue
                
                # If we've already successfully matched a size for this specific column name in this row, skip other types for it
                if col in processed_columns_in_row:
                    continue

                val = str(row.get(col))
                if val and val != 'None' and val.strip():
                    # Try to find size linked to product tables first
                    matching_size = Size.objects.filter(
                        table__products=product,
                        type=s_type,
                        value=val
                    ).first()
                    
                    if not matching_size:
                        # Search globally for this size
                        matching_size = Size.objects.filter(type=s_type, value=val).first()
                        
                        if not matching_size:
                            # Not found globally? Create it in a Local SizeTable.
                            brand = product.brand
                            size_table, _ = SizeTable.objects.get_or_create(brand=brand, name="Local Sizes")
                            matching_size = Size.objects.create(
                                table=size_table,
                                type=s_type,
                                value=val,
                                insole_min_mm=0, 
                                insole_max_mm=0
                            )
                        
                        # Since it was a global/local fall-back, mark the partner product as local-only
                        if not partner_product.online: # already local?
                            pass
                        else:
                            partner_product.online = False
                            partner_product.save()

                    if matching_size:
                        processed_columns_in_row.add(col) # Mark this column as "done" for this row
                        if quantity > 0:
                            pps, pps_created = PartnerProductSize.objects.get_or_create(
                                partner_product=partner_product,
                                size=matching_size,
                                defaults={'quantity': quantity}
                            )
                            if not pps_created:
                                # Add to existing quantity instead of overwriting
                                pps.quantity += quantity 
                                pps.save()
                            summary["size_updates"] += 1

        return Response({
            "message": "File processed with local-only fallback logic",
            "summary": summary,
            "debug": {
                "skipped_details": skipped_details[:50] # Limit debug info
            }
        }, status=status.HTTP_200_OK)


    def clean_price(self, price_str):
        if not price_str: return Decimal('0.00')
        # Remove currency symbols and spaces
        cleaned = "".join(c for c in str(price_str) if c.isdigit() or c in '.,')
        if not cleaned: return Decimal('0.00')
        
        # Consistent decimal separator
        if ',' in cleaned and '.' in cleaned:
            cleaned = cleaned.replace(',', '')
        elif ',' in cleaned:
            cleaned = cleaned.replace(',', '.')
        
        try:
            return Decimal(cleaned)
        except:
            return Decimal('0.00')


class AddLocalOnlyPartnerProduct(views.APIView):
    permission_classes = [permissions.IsAuthenticated, IsPartner]

    def post(self, request):
        try:
            name = request.data.get('name')
            brand_name = request.data.get('brand')
            price = request.data.get('price', 0)
            colors = request.data.get('colors') or request.data.get('colos') or request.data.get('color') or []
            warehouse_id = request.data.get('warehouse')
            sizes_input = request.data.get('sizes', []) # Expecting list/dict like {"EU 40": 20}
            eanc = request.data.get('eanc', None)
            buy_price = request.data.get('buy_price', 0)
            
            if not name:
                return Response({"error": "Name is required."}, status=status.HTTP_400_BAD_REQUEST)

            # 1. Resolve Brand
            brand = Brand.objects.filter(name__iexact=brand_name).first()
            if not brand:
                brand = Brand.objects.create(name=brand_name)

            # 2. Resolve or Create Product
            product, created = Product.objects.get_or_create(
                name=name,
                brand=brand,
                defaults={
                    'gender': 'unisex',
                    'description': ' ',
                    'is_active': False
                }
            )

            # 3. Resolve Color
            color_input = request.data.get('color')
            if not color_input:
                return Response({"error": "Color is required."}, status=status.HTTP_400_BAD_REQUEST)
            
            color_input = str(color_input).strip()
            color_obj = Color.objects.filter(color__iexact=color_input).first()
            if not color_obj:
                color_obj = Color.objects.create(color=color_input, hex_code="#CCCCCC")
            
            # 4. Create single PartnerProduct variant
            partner_product, pp_created = PartnerProduct.objects.get_or_create(
                partner=request.user,
                product=product,
                color=color_obj,
                defaults={
                    'price': price,
                    'local': True,
                    'online': False,
                    'eanc': eanc,
                    'buy_price': buy_price,
                }
            )

            if not pp_created:
                partner_product.price = price
                if eanc:
                    partner_product.eanc = eanc
                partner_product.save()

            if warehouse_id:
                from Others.models import Warehouse
                warehouse = Warehouse.objects.filter(id=warehouse_id, partner=request.user).first()
                if warehouse:
                    warehouse.product.add(partner_product)

            # 4. Handle sizes and quantities for this specific variant
            if isinstance(sizes_input, dict):
                sizes_items = sizes_input.items()
            elif isinstance(sizes_input, list):
                sizes_items = []
                for item in sizes_input:
                    if isinstance(item, dict):
                        sizes_items.extend(item.items())
            else:
                sizes_items = []

            if sizes_items:
                size_table, _ = SizeTable.objects.get_or_create(brand=brand, name=f"{brand.name} Sizes")
                
                for size_str, quantity in sizes_items:
                    parts = str(size_str).split()
                    if len(parts) >= 2:
                        s_type = parts[0]
                        s_value = " ".join(parts[1:])
                    else:
                        s_type = "EU"
                        s_value = size_str

                    size_obj, _ = Size.objects.get_or_create(
                        table=size_table,
                        type=s_type,
                        value=s_value,
                        defaults={'insole_min_mm': 0, 'insole_max_mm': 0}
                    )
                    
                    try:
                        qty = int(quantity) if quantity else 0
                    except (ValueError, TypeError):
                        qty = 0

                    pps, pps_created = PartnerProductSize.objects.update_or_create(
                        partner_product=partner_product,
                        size=size_obj,
                        defaults={'quantity': qty}
                    )

            return Response({"message": "Product variant created successfully", "id": partner_product.id}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": f"Failed to add local product: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
